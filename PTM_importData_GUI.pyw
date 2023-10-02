# Tanunnas BK ©2023
#region import
import os, sys, io, openpyxl
from tkinter import Tk, Menu, messagebox, filedialog, Listbox, Scrollbar, ttk
from PyPDF2 import PdfReader
from openpyxl.styles import PatternFill, Border, Side
#endregion import

#region special function
#ตรวจสอบว่าเป็น path ปกติ หรือ resource_path ของ PyInstaller
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)
#endregion special function

#region FormInit
mainfrm = Tk()
mainfrm.geometry("315x350+50+150")
mainfrm.resizable(width=False, height=False)
mainfrm.title("PTM - Data import from Pdf 1.0")
mainfrm.wm_iconbitmap(resource_path("ptm_icon.ico"))
#endregion FormInit

#region variable
curr_directory = os.getcwd()
fileName = ""
data = []
#endregion variable

#region Main function
def openPDF():
    global data, fileName
    
    fileName = filedialog.askopenfilename(title="Open Pdf..", initialdir=curr_directory, filetypes=(('Pdf files','*.pdf'),))     
    if fileName != "":
        listData.delete(0,"end")
        reader = PdfReader(fileName)      
                
        for page in reader.pages:    
            page_text = page.extract_text()    
            ref1 = name = address = province = postcode = ""
            
            for line in io.StringIO(page_text):
                    if "Ref1" in line:            
                        ref1 = line.split(": ")
                        ref1 = ref1[1].strip()                    
                    elif "นามสกุล :" in line:
                        name = line.split(": ")
                        name = name[1].strip()                    
                    elif "ที่อยู่ :" in line:
                        address = line.split(": ")
                        address = address[1].strip()                    
                    elif "อ." in line and line.find("อ.") < 3:  #พบในตำแหน่งต้นๆ ของบรรทัด          
                        province = line.strip()                    
                    elif "ไปรษณีย์ :" in line:                
                        postcode = line.split(": ")
                        postcode = postcode[1].strip()                    

            row = [ref1, name, address, province, postcode]    
            data.append(row)  
            listData.insert('end', ref1 + "  " + name)        
        bntSave.config(state="enabled") 
        messagebox.showinfo("Import data", "นำเข้าข้อมูลเรียบร้อย.")

def saveToExcel():
    global fileName
    header = ['Ref1', 'Name', 'Address', 'Province', 'Postcode']
    if len(fileName) == 0:
        messagebox.showinfo("Error", "กรุณาเปิดไฟล์ Pdf เพื่อนำเข้าข้อมูลก่อน!")
        return

    fileName = fileName.split(".")[0] + ".xlsx"

    wb = openpyxl.Workbook()
    sheet = wb.worksheets[0]
    sheet.append(header)
    for row in data :
        sheet.append(row)

    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['E'].width = 10

    thin = Side(border_style="thin", color="000000")

    for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=None): 
        for cell in rows: 
            cell.fill = PatternFill(start_color="80FF00", end_color="80FF00",fill_type = "solid")
            cell.border = Border(top=None, left=None, right=None, bottom=thin)
    wb.save(fileName)
    wb.close()
    messagebox.showinfo("Save Data", "บันทึกไฟล์เรียบร้อย.")

def msgAbout():    
    messagebox.showinfo("เกี่ยวกับโปรแกรม", "โปรแกรม : PTM - Data Import from Pdf v1.0\n ผู้พัฒนา : Tanunnas BK")

#endregion Main function

#region GUI
# Menubar
menubar = Menu(mainfrm)
file = Menu(menubar, tearoff=0)
file.add_command(label="Open Pdf..", command=openPDF)
file.add_command(label="Save to Excel..", command=saveToExcel)

file.add_separator()
file.add_command(label="Exit", command=mainfrm.quit)
menubar.add_cascade(label="File", menu=file)

menubar.add_cascade(label="About", command=msgAbout)
mainfrm.config(menu=menubar)

lblfrm = ttk.LabelFrame(mainfrm, text="รายการข้อมูล")
lblfrm.grid(column=0, row=0, columnspan=2, padx=10, pady=5, sticky="w")
listData = Listbox(lblfrm, name="listData", height=12, width=45)
listData.pack(side="left", fill="y", pady=2)

scrollbar1 = Scrollbar(lblfrm, orient="vertical" )
scrollbar1.config(command=listData.yview)            #เชื่อมกับ listbox
scrollbar1.pack(side="left", fill="y", pady=2)

listData.config(yscrollcommand=scrollbar1.set)       #เชื่อมกับ scrollbar

lblfrm3 = ttk.LabelFrame(mainfrm, text="การบันทึกข้อมูล")
lblfrm3.grid(column=0, row=2, columnspan=2, padx=10, pady=0, sticky="w")
bntReset = ttk.Button(lblfrm3, text="เปิดไฟล์ Pdf", command=openPDF)
bntReset.grid(column=0, row=1, padx=10, pady=10)
bntSave = ttk.Button(lblfrm3, text="บันทึกเป็น Excel", command=saveToExcel)
bntSave.grid(column=1, row=1, padx=10, pady=10)
bntSave.config(state="disabled")  #ปิดการใช้งานปุ่มเซฟ

#endregion GUI

mainfrm.mainloop()
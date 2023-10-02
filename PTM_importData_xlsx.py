# Tanunnas BK ©2023
import os, io, subprocess, openpyxl
from PyPDF2 import PdfReader
from openpyxl.styles import PatternFill, Border, Side

if input("ต้องการเปิดโฟลเดอร์ข้อมูลหรือไม่ (1:ใช่) : ") == "1":
    subprocess.Popen(r'explorer /select,' + __file__)

while 1:
    fileName = input("ป้อนชื่อไฟล์หนังสือแจ้งเตือน (.pdf) : ")
    if fileName.find(".") == -1:
        fileName = fileName + ".pdf"
    elif fileName.split(".")[1] != "pdf":
        print("ผิดพลาด : ต้องเป็นไฟล์ .pdf เท่านั้น !")
        continue        
    try:
        reader = PdfReader(fileName)        
        break
    except FileNotFoundError:
        print("ผิดพลาด : ชื่อไฟล์ไม่ถูกต้อง !")
        
pageNum = len(reader.pages)
header = ['Ref1', 'Name', 'Address', 'Province', 'Postcode']
data = []
os.system('cls')
print("\033[3;37;42m   Ref1           ชื่อ - สกุล                        ที่อยู่                       อำเภอ/จังหวัด           รหัสไปรษณีย์ \033[0;37;40m")

for page in reader.pages:    
    page_text = page.extract_text()    
    ref1 = name = address = province = postcode = ""
    
    for line in io.StringIO(page_text):
            if "Ref1" in line:            
                ref1 = line.split(": ")
                ref1 = ref1[1].strip()
                print("{0:<15}".format(ref1), end="")
            elif "นามสกุล :" in line:
                name = line.split(": ")
                name = name[1].strip()
                print("{0:<35}".format(name[0:33]), end="")
            elif "ที่อยู่ :" in line:
                address = line.split(": ")
                address = address[1].strip()
                print("{0:<30}".format(address[0:28]), end="")
            elif "อ." in line and line.find("อ.") < 3:  #พบในตำแหน่งต้นๆ ของบรรทัด          
                province = line.strip()
                print("{0:<30}".format(province[0:28]), end="")
            elif "ไปรษณีย์ :" in line:                
                postcode = line.split(": ")
                postcode = postcode[1].strip()
                print("{0:<9}".format(postcode)) 

    row = [ref1, name, address, province, postcode]    
    data.append(row)          

print("-" * 120)
print("รวมหนังสือแจ้ง จำนวน : ", pageNum)
name = input("\nตั้งชื่อไฟล์ข้อมูลที่จะบันทึก (.xlsx) : ")
if len(name)!= 0:       
    if "." in name:        
        fileName = name.split(".")[0] + ".xlsx"
    else: fileName = name + ".xlsx"
else:
    fileName = fileName.split(".")[0] + ".xlsx"

wb = openpyxl.Workbook()
sheet = wb.worksheets[0]
sheet.append(header)
for row in data :
    sheet.append(row)

sheet.column_dimensions['A'].width = 15
sheet.column_dimensions['B'].width = 40
sheet.column_dimensions['C'].width = 25
sheet.column_dimensions['D'].width = 25
sheet.column_dimensions['E'].width = 10

thin = Side(border_style="thin", color="000000")

for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=None): 
   for cell in rows: 
        cell.fill = PatternFill(start_color="80FF00", end_color="80FF00",fill_type = "solid")
        cell.border = Border(top=None, left=None, right=None, bottom=thin)
wb.save(fileName)
wb.close()
         
input("\nบันทึกเรียบร้อย กดปุ่ม \"Enter\" เพื่อปิดโปรแกรม..")